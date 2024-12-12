###First Replit

# import openpyxl
# import json
# import requests


# url='https://api.coincap.io/v2/assets'

# response=requests.get(url)
# print(response)

# data=response.json()

# for cur in data ["data"]:
#   print(cur["name"],cur["symbol"],"-rank", cur["rank"])

# ##Second Replit
# import openpyxl

# from openpyxl.utils import get_column_letter,column_index_from_string
# wb = openpyxl.load_workbook('Students-BTA-Connected-Data.xlsx')

# print(get_column_letter(1))
# print(get_column_letter(27))
# print(get_column_letter(1212))
# print(get_column_letter(9999))

# sheet = wb['Sheet1']
# print(get_column_letter(sheet.max_column))

# print(column_index_from_string('A'))
# print(column_index_from_string('AA'))
# print(column_index_from_string('ATP'))
# print(column_index_from_string('NTO'))

# ###Third Replit

# import openpyxl
# wb = openpyxl.load_workbook('Students-BTA-Connected-Data.xlsx')
# sheet = wb['Sheet1']

# print(type(sheet)) #the type of the data
# print(wb.sheetnames) #the names of the sheets
# print(sheet)
# print(type(sheet))
# print(sheet.title) #prints the title of the sheet
# print(sheet['A1'].value) #value of location A1
# print(sheet['B1'].value) #value of location B1
# print(sheet['C1'].value) #value of location C1

# name = sheet['B1']

# print(f"Row{name.row}, Column{name.column}, Value is {name.value}")

# print(sheet.cell(row=1, column=2))
# print(sheet.cell(row=1, column=2).value)

# for value in range(1,5,1):
#   #print(value)
#   print(value, sheet.cell(row=value,column=1).value,
#        sheet.cell(row=value,column=2).value,
#        sheet.cell(row=value,column=3).value
#        )

# #Shape of the sheet
# print(sheet.max_row)
# print(sheet.max_column)


# ##Fourth Replit

# from typing import dataclass_transform
# import openpyxl
# wb = openpyxl.load_workbook('Students-BTA-Connected-Data.xlsx')
# sheet = wb['Sheet1']

# data = tuple(sheet['A1':'C4']) #Get all the cell from A1 till C4

# print(data)

# #print rows and columns

# for row in sheet['A1':'C4']:
#   for cell in row:
#     print(cell.coordinate, cell.value)
#   print("END of the ROW".center(30,"-"))


# ##fifth Replit

# import openpyxl

# wb=openpyxl.Workbook() #this is virtual workbook

# sheet=wb.active

# #We create our data / you can get from api

# for value in range(1,11,1):
#   sheet['A'+str(value)]=value #A1 -1, A2 -2 ....

# refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1,
# max_row=10)

# seriesObj = openpyxl.chart.Series(refObj, title='Monthly Sales Report')

# chartObj = openpyxl.chart.BarChart()  #the type of the chart

# chartObj.title = 'My Chart'
# chartObj.append(seriesObj)

# sheet.add_chart(chartObj, 'C5')
# wb.save('Excel/sampleChart.xlsx')
# print('Done')




import openpyxl
from openpyxl.chart import BarChart, Reference

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active

# dataset with names and ages
ages = {
    "Emml": 37,
    "Mike": 34,
    "Kwesi": 36,
    "Nana": 40,
    "Asma": 42
}

# Add names and ages to the sheet
sheet['A1'] = 'Name'
sheet['B1'] = 'Age'
for row, (name, age) in enumerate(ages.items(), start=2):
    sheet[f'A{row}'] = name
    sheet[f'B{row}'] = age

# Calculate the mean age
age_values = list(ages.values())
mean_age = sum(age_values) / len(age_values)

# Add the mean age to the sheet
sheet['D1'] = 'Mean Age'
sheet['D2'] = mean_age

# Create a reference for the ages data
refObj = Reference(sheet, min_col=2, min_row=2, max_col=2, max_row=len(ages) + 1)

# Create a series and add it to the chart
chartObj = BarChart()
chartObj.title = 'Age Chart'
chartObj.x_axis.title = 'Names'
chartObj.y_axis.title = 'Ages'
chartObj.add_data(refObj, titles_from_data=False)

# Add categories (names) to the chart
chartObj.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=len(ages) + 1))

# Add the chart to the sheet at position D5
sheet.add_chart(chartObj, 'D5')

# Save the workbook
wb.save('Excel/AgeChart.xlsx')
print('Done')
